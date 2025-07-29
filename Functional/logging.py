from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime



class LogApp:
    def __init__(self):
        now = datetime.now()
        timestamp = now.strftime("%d-%m-%Y--%H-%M")  # e.g. "20250729-142530"
        self.filename = f"logs_{timestamp}.xlsx"
        self.OUTPUT0_PATH = Path(__file__).parent
        self.OUTPUT1_PATH = self.OUTPUT0_PATH.parent
        self.LOG_FILE_PATH = self.OUTPUT1_PATH / "Logs" / self.filename
        
        self.LOG_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
        
        if not self.LOG_FILE_PATH.is_file():
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            # Write header row
            ws.append(["Sr No","DID", "Result"])
            wb.save(self.LOG_FILE_PATH)
            print(f"Created new Excel file with header at: {self.LOG_FILE_PATH}")
        else:
            print(f"Excel file already exists at: {self.LOG_FILE_PATH}")
        
        # Store workbook and worksheet as instance attributes
        self.wb = load_workbook(self.LOG_FILE_PATH)
        self.ws = self.wb.active
        self.sr_no = 1
        
    def add_log(self, did: str, result: str):
        self.ws.append([self.sr_no, did, result])
        self.wb.save(self.LOG_FILE_PATH)
        self.sr_no += 1
